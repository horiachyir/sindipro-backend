from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from .models import Building, Tower, Unit
from .serializers import BuildingSerializer, BuildingReadSerializer, UnitSerializer, UnitDetailSerializer, BuildingBasicSerializer
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import openpyxl
import io

def user_can_access_building(user, building_id):
    """
    Check if user can access a specific building.
    Master role users can access all buildings.
    Other users can only access buildings they created.
    """
    if user.role == 'master':
        return True

    try:
        Building.objects.get(id=building_id, created_by=user)
        return True
    except Building.DoesNotExist:
        return False

def get_accessible_building(user, building_id):
    """
    Get building if user has access, raise appropriate exception if not.
    """
    if user.role == 'master':
        try:
            return Building.objects.get(id=building_id)
        except Building.DoesNotExist:
            return None
    else:
        try:
            return Building.objects.get(id=building_id, created_by=user)
        except Building.DoesNotExist:
            return None

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def get_buildings(request):
    if request.method == 'GET':
        # Master role users can see all buildings, others only see buildings they created
        if request.user.role == 'master':
            buildings = Building.objects.all().prefetch_related('address', 'alternative_address', 'towers__unit_distribution')
        else:
            buildings = Building.objects.filter(created_by=request.user).prefetch_related('address', 'alternative_address', 'towers__unit_distribution')
        serializer = BuildingReadSerializer(buildings, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'POST':
        serializer = BuildingSerializer(data=request.data)

        if serializer.is_valid():
            building = serializer.save(created_by=request.user)
            return Response({
                'message': 'Building created successfully',
                'building_id': building.id,
                'building_name': building.building_name
            }, status=status.HTTP_201_CREATED)

        return Response({
            'error': 'Invalid data',
            'details': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_building(request):
    serializer = BuildingSerializer(data=request.data)

    if serializer.is_valid():
        building = serializer.save(created_by=request.user)
        return Response({
            'message': 'Building created successfully',
            'building_id': building.id,
            'building_name': building.building_name
        }, status=status.HTTP_201_CREATED)

    return Response({
        'error': 'Invalid data',
        'details': serializer.errors
    }, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def update_building(request, id):
    # Check if user has access to building (master role can access all buildings)
    building = get_accessible_building(request.user, id)
    if not building:
        return Response({
            'error': 'Building not found or access denied'
        }, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'DELETE':
        # Store building name for response message
        building_name = building.building_name

        # Explicitly delete address records from building_mgmt_address table
        if building.address:
            building.address.delete()

        if building.alternative_address:
            building.alternative_address.delete()

        # Delete the building (this will cascade delete all related records including towers)
        building.delete()

        return Response({
            'message': f'Building "{building_name}" and all associated data have been successfully deleted'
        }, status=status.HTTP_200_OK)

    elif request.method == 'PUT':
        serializer = BuildingSerializer(building, data=request.data)

        if serializer.is_valid():
            updated_building = serializer.save()
            return Response({
                'message': 'Building updated successfully',
                'building_id': updated_building.id,
                'building_name': updated_building.building_name
            }, status=status.HTTP_200_OK)

        return Response({
            'error': 'Invalid data',
            'details': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_all_buildings(request):
    """
    Get all buildings with all fields from building_mgmt_building, building_mgmt_address, and building_mgmt_tower tables.
    Accessible without authentication for frontend signup process.
    """
    buildings = Building.objects.all().select_related('address', 'alternative_address').prefetch_related('towers__unit_distribution')
    serializer = BuildingReadSerializer(buildings, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_units(request):
    # Master role users can see all units, others only see units from their buildings
    if request.user.role == 'master':
        units = Unit.objects.all().select_related('building', 'tower').order_by('building__building_name', 'number')
    else:
        units = Unit.objects.filter(
            building__created_by=request.user
        ).select_related('building', 'tower').order_by('building__building_name', 'number')

    serializer = UnitDetailSerializer(units, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_unit(request, id):
    print(f"DEBUG: Received request for building_id: {id}")
    print(f"DEBUG: User: {request.user}")
    print(f"DEBUG: Request data: {request.data}")

    # Verify the building exists and user has access (master role can access all buildings)
    building = get_accessible_building(request.user, id)
    if not building:
        print(f"DEBUG: Building {id} not found or access denied for user {request.user}")
        return Response({
            'error': 'Building not found or access denied'
        }, status=status.HTTP_404_NOT_FOUND)

    print(f"DEBUG: Found building: {building}")

    # Remove building_id from request data if present (since it comes from URL)
    data = request.data.copy()
    if 'building_id' in data:
        data.pop('building_id')

    serializer = UnitSerializer(data=data)

    if serializer.is_valid():
        # Set the building from the URL parameter
        unit = serializer.save(building=building)
        return Response({
            'message': 'Unit created successfully',
            'unit': UnitSerializer(unit).data
        }, status=status.HTTP_201_CREATED)

    return Response({
        'error': 'Invalid data',
        'details': serializer.errors
    }, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def update_unit(request, id):
    # Add debug information
    print(f"DEBUG: update_unit called with ID: {id}")
    print(f"DEBUG: Request method: {request.method}")
    print(f"DEBUG: User: {request.user}")

    try:
        # First check if unit exists at all
        unit = Unit.objects.get(id=id)
        print(f"DEBUG: Found unit: {unit.number} in building: {unit.building.building_name}")
        print(f"DEBUG: Building created by: {unit.building.created_by}")
    except Unit.DoesNotExist:
        print(f"DEBUG: Unit with ID {id} not found in database")
        return Response({
            'error': f'Unit with ID {id} not found'
        }, status=status.HTTP_404_NOT_FOUND)

    # Then check if user has access to this unit's building (master role can access all buildings)
    if not user_can_access_building(request.user, unit.building.id):
        print(f"DEBUG: Access denied. Building owner: {unit.building.created_by}, Request user: {request.user}")
        return Response({
            'error': 'Access denied. You can only modify units in buildings you created.',
            'debug_info': {
                'building_owner': str(unit.building.created_by),
                'request_user': str(request.user),
                'unit_id': id,
                'building_name': unit.building.building_name
            }
        }, status=status.HTTP_403_FORBIDDEN)

    if request.method == 'DELETE':
        # Store unit info for response message
        unit_number = unit.number
        building_name = unit.building.building_name

        # Delete the unit from building_mgmt_unit table
        unit.delete()

        return Response({
            'message': f'Unit {unit_number} from building "{building_name}" has been successfully deleted'
        }, status=status.HTTP_200_OK)

    elif request.method == 'PUT':
        serializer = UnitSerializer(unit, data=request.data)

        if serializer.is_valid():
            updated_unit = serializer.save()
            return Response({
                'message': 'Unit updated successfully',
                'unit': UnitDetailSerializer(updated_unit).data
            }, status=status.HTTP_200_OK)

        return Response({
            'error': 'Invalid data',
            'details': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_units_excel(request, id):
    # Check if user has access to building (master role can access all buildings)
    building = get_accessible_building(request.user, id)
    if not building:
        return Response({
            'error': 'Building not found or access denied'
        }, status=status.HTTP_404_NOT_FOUND)

    # Get all units for this building with related data
    units = Unit.objects.filter(building=building).select_related(
        'building', 'tower'
    ).order_by('tower__name', 'floor', 'number')

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = f"{building.building_name} - Units Report"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Title and building information
    ws.merge_cells('A1:L1')
    title_cell = ws['A1']
    title_cell.value = f"Building Units Report - {building.building_name}"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells('A2:L2')
    building_info = ws['A2']
    address = building.address
    address_str = f"{address.street}, {address.number}, {address.neighborhood}, {address.city}/{address.state}"
    building_info.value = f"Address: {address_str} | CNPJ: {building.cnpj} | Manager: {building.manager_name}"
    building_info.alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        'Unit Number', 'Tower', 'Floor', 'Identification', 'Area (m²)',
        'Ideal Fraction', 'Status', 'Owner', 'Owner Phone', 'Parking Spaces',
        'Key Delivery', 'Deposit Location'
    ]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Write unit data
    row = 5
    for unit in units:
        # Display names instead of IDs
        tower_name = unit.tower.name if unit.tower else "N/A"

        # Map status values to readable text
        status_display = dict(Unit.STATUS_CHOICES).get(unit.status, unit.status)

        # Map identification values to readable text
        identification_display = dict(Unit.IDENTIFICATION_CHOICES).get(unit.identification, unit.identification)

        # Map key_delivery to more readable text for export
        key_delivery_display_map = {
            'Yes': 'Yes',
            'No': 'No',
            'Pnd': 'Pending',
            'Y': 'Yes',
            'N': 'No'
        }
        key_delivery_display = key_delivery_display_map.get(unit.key_delivery, unit.key_delivery)

        data = [
            unit.number,
            tower_name,
            unit.floor,
            identification_display,
            float(unit.area),
            float(unit.ideal_fraction),
            status_display,
            unit.owner,
            unit.owner_phone,
            unit.parking_spaces,
            key_delivery_display,
            unit.deposit_location or "N/A"
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = border
            # Center alignment for numeric and status fields
            if col in [3, 4, 5, 6, 7, 10]:  # Floor, identification, area, fraction, status, parking
                cell.alignment = Alignment(horizontal="center")

        row += 1

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0

        # Check header length
        max_length = max(max_length, len(str(headers[col-1])))

        # Check data lengths
        for row_num in range(5, row):
            cell_value = ws.cell(row=row_num, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))

        # Set column width with some padding
        adjusted_width = min(max_length + 2, 50)  # Max width of 50
        ws.column_dimensions[column_letter].width = adjusted_width

    # Add summary information
    summary_row = row + 2
    ws.merge_cells(f'A{summary_row}:D{summary_row}')
    summary_cell = ws[f'A{summary_row}']
    summary_cell.value = f"Total Units: {units.count()}"
    summary_cell.font = Font(bold=True)

    # Status summary
    status_counts = {}
    for unit in units:
        status_display = dict(Unit.STATUS_CHOICES).get(unit.status, unit.status)
        status_counts[status_display] = status_counts.get(status_display, 0) + 1

    summary_row += 1
    for status_name, count in status_counts.items():
        ws.merge_cells(f'A{summary_row}:D{summary_row}')
        status_cell = ws[f'A{summary_row}']
        status_cell.value = f"{status_name}: {count} units"
        summary_row += 1

    # Create HTTP response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    filename = f"{building.building_name.replace(' ', '_')}_units_report.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_units_excel(request, id):
    try:
        # Add comprehensive logging at the very start
        print(f"DEBUG: Excel import started - Building ID: {id}, User: {request.user}")
        print(f"DEBUG: Request method: {request.method}")
        print(f"DEBUG: Request FILES keys: {list(request.FILES.keys())}")
        print(f"DEBUG: Request DATA keys: {list(request.data.keys())}")
    except Exception as debug_error:
        print(f"DEBUG: Error in initial logging: {str(debug_error)}")
        return Response({
            'error': f'Error in initial request processing: {str(debug_error)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    try:
        print(f"DEBUG: Attempting to find building with ID {id}")
        # Check if user has access to building (master role can access all buildings)
        building = get_accessible_building(request.user, id)
        if not building:
            print(f"DEBUG: Building {id} not found or access denied for user {request.user}")
            return Response({
                'error': 'Building not found or access denied'
            }, status=status.HTTP_404_NOT_FOUND)
        print(f"DEBUG: Found building: {building.building_name}")
    except Exception as e:
        print(f"DEBUG: Database error while finding building: {str(e)}")
        return Response({
            'error': f'Database error: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # Check if file was uploaded
    if 'file' not in request.FILES:
        return Response({
            'error': 'No file uploaded. Please upload an Excel file.'
        }, status=status.HTTP_400_BAD_REQUEST)

    excel_file = request.FILES['file']

    # Validate file type
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        return Response({
            'error': 'Invalid file type. Please upload an Excel file (.xlsx or .xls).'
        }, status=status.HTTP_400_BAD_REQUEST)

    try:
        # Load workbook
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
    except Exception as e:
        return Response({
            'error': f'Error reading Excel file: {str(e)}'
        }, status=status.HTTP_400_BAD_REQUEST)

    try:
        # Find header row (should be row 4 based on export format)
        header_row = 4
        headers = []
        for col in range(1, 13):  # 12 columns expected
            try:
                cell_value = ws.cell(row=header_row, column=col).value
                if cell_value:
                    headers.append(str(cell_value).strip())
                else:
                    headers.append("")
            except Exception:
                headers.append("")

        # Expected headers from export
        expected_headers = [
            'Unit Number', 'Tower', 'Floor', 'Identification', 'Area (m²)',
            'Ideal Fraction', 'Status', 'Owner', 'Owner Phone', 'Parking Spaces',
            'Key Delivery', 'Deposit Location'
        ]

        # Validate headers (more flexible validation)
        if len(headers) < 8:  # At least basic headers
            return Response({
                'error': 'Invalid Excel format. Missing required columns.',
                'expected_headers': expected_headers,
                'found_headers': headers
            }, status=status.HTTP_400_BAD_REQUEST)

        # Parse data rows (starting from row 5)
        units_data = []
        errors = []
        row_num = 5

        # Get all towers for this building to map tower names to IDs
        try:
            towers = {tower.name: tower for tower in building.towers.all()}
        except Exception as e:
            return Response({
                'error': f'Error loading building towers: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Create reverse mappings for choices
        status_map = {v: k for k, v in Unit.STATUS_CHOICES}
        identification_map = {v: k for k, v in Unit.IDENTIFICATION_CHOICES}

        # Process up to 1000 rows to prevent infinite loops
        max_rows = 1000
        processed_rows = 0

        while processed_rows < max_rows:
            try:
                # Check if row has data (check first column)
                unit_number_cell = ws.cell(row=row_num, column=1).value
                if not unit_number_cell or str(unit_number_cell).strip() == '':
                    break

                # Extract all cell values safely
                row_data = []
                for col in range(1, 13):
                    try:
                        cell_value = ws.cell(row=row_num, column=col).value
                        row_data.append(cell_value)
                    except Exception:
                        row_data.append(None)

                # Parse and validate each field with safe conversion
                try:
                    # Safe string conversion with encoding handling
                    def safe_str(value, default=''):
                        if value is None:
                            return default
                        try:
                            # Convert to string and ensure it's properly encoded
                            s = str(value).strip()
                            # Encode and decode to ensure valid UTF-8
                            return s.encode('utf-8', errors='ignore').decode('utf-8', errors='ignore')
                        except Exception:
                            return default

                    unit_number = safe_str(row_data[0], '')
                    tower_name_raw = safe_str(row_data[1], '')
                    tower_name = tower_name_raw if tower_name_raw and tower_name_raw != 'N/A' else None

                    # Handle floor conversion
                    try:
                        floor = int(float(row_data[2])) if row_data[2] is not None else 1
                    except (ValueError, TypeError):
                        floor = 1

                    identification_display = safe_str(row_data[3], 'Residential')

                    # Handle area conversion
                    try:
                        area = float(row_data[4]) if row_data[4] is not None else 0.0
                    except (ValueError, TypeError):
                        area = 0.0

                    # Handle ideal_fraction conversion
                    try:
                        ideal_fraction = float(row_data[5]) if row_data[5] is not None else 0.0
                    except (ValueError, TypeError):
                        ideal_fraction = 0.0

                    status_display = safe_str(row_data[6], 'Vacant')
                    owner = safe_str(row_data[7], '')
                    owner_phone = safe_str(row_data[8], '')

                    # Handle parking_spaces conversion
                    try:
                        parking_spaces = int(float(row_data[9])) if row_data[9] is not None else 0
                    except (ValueError, TypeError):
                        parking_spaces = 0

                    key_delivery_input = safe_str(row_data[10], 'No')
                    deposit_location = safe_str(row_data[11], '')

                    # Map display values back to database values
                    unit_status = status_map.get(status_display, 'vacant')
                    identification = identification_map.get(identification_display, 'residential')

                    # Map key_delivery values to database-compatible short codes
                    key_delivery_map = {
                        'yes': 'Yes',
                        'no': 'No',
                        'delivered': 'Yes',
                        'not delivered': 'No',
                        'pending': 'Pnd',
                        'received': 'Yes',
                        'given': 'Yes',
                        'done': 'Yes',
                        'not done': 'No',
                        'completed': 'Yes',
                        'incomplete': 'No',
                        'sim': 'Yes',  # Portuguese for yes
                        'não': 'No',   # Portuguese for no
                        'nao': 'No',   # Portuguese for no (without accent)
                        'entregue': 'Yes',  # Portuguese for delivered
                    }

                    # Normalize and map key_delivery value
                    key_delivery_normalized = key_delivery_input.lower().strip()
                    key_delivery = key_delivery_map.get(key_delivery_normalized, 'No')

                    # If not found in map and has a value, try to truncate safely
                    if key_delivery == 'No' and key_delivery_input:
                        # Ensure the truncation doesn't break UTF-8 encoding
                        truncated = key_delivery_input[:3]
                        # Verify it's valid UTF-8 after truncation
                        try:
                            truncated.encode('utf-8')
                            key_delivery = truncated
                        except UnicodeEncodeError:
                            key_delivery = 'No'

                    # Validate and truncate fields to prevent database errors
                    # Ensure truncation doesn't break UTF-8 encoding
                    unit_number = unit_number[:20] if len(unit_number) > 20 else unit_number
                    owner = owner[:200] if len(owner) > 200 else owner
                    owner_phone = owner_phone[:20] if len(owner_phone) > 20 else owner_phone
                    deposit_location = deposit_location[:200] if len(deposit_location) > 200 else deposit_location

                    # Ensure identification and status are within limits
                    identification = identification[:20] if len(identification) > 20 else identification
                    unit_status = unit_status[:20] if len(unit_status) > 20 else unit_status
                    key_delivery = key_delivery[:3] if len(key_delivery) > 3 else key_delivery

                    # Find tower by name
                    tower = None
                    if tower_name:
                        tower = towers.get(tower_name)
                        if not tower:
                            errors.append(f"Row {row_num}: Tower '{tower_name}' not found in building")
                            row_num += 1
                            processed_rows += 1
                            continue

                    # Validate required fields
                    if not unit_number:
                        errors.append(f"Row {row_num}: Unit number is required")
                        row_num += 1
                        processed_rows += 1
                        continue

                    if area <= 0:
                        errors.append(f"Row {row_num}: Area must be greater than 0")
                        row_num += 1
                        processed_rows += 1
                        continue

                    # Create unit data
                    unit_data = {
                        'building': building,
                        'tower': tower,
                        'number': unit_number,
                        'floor': floor,
                        'area': area,
                        'ideal_fraction': ideal_fraction,
                        'identification': identification,
                        'deposit_location': deposit_location,
                        'key_delivery': key_delivery,
                        'owner': owner,
                        'owner_phone': owner_phone,
                        'parking_spaces': parking_spaces,
                        'status': unit_status,
                    }

                    units_data.append(unit_data)

                except Exception as e:
                    errors.append(f"Row {row_num}: Error parsing data - {str(e)}")

            except Exception as e:
                errors.append(f"Row {row_num}: Error reading row - {str(e)}")

            row_num += 1
            processed_rows += 1

        # If there are validation errors, return them
        if errors and not units_data:
            return Response({
                'error': 'Data validation failed',
                'details': errors
            }, status=status.HTTP_400_BAD_REQUEST)

        if not units_data:
            return Response({
                'error': 'No valid unit data found in the Excel file'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Save units to database with transaction and batch processing
        created_units = []
        update_count = 0
        create_count = 0
        save_errors = []

        print(f"DEBUG: Starting to save {len(units_data)} units to database")

        # Pre-fetch all existing units for this building to avoid repeated queries
        # This prevents timeout issues from querying potentially corrupted data
        try:
            existing_units_dict = {}
            # Get all units for this building at once
            existing_units = Unit.objects.filter(building_id=building.id).values('id', 'number')
            for unit in existing_units:
                try:
                    # Safely decode the number field
                    number = unit['number']
                    if isinstance(number, bytes):
                        number = number.decode('utf-8', errors='ignore')
                    existing_units_dict[number] = unit['id']
                except Exception as decode_error:
                    print(f"DEBUG: Error decoding existing unit: {decode_error}")
                    continue
            print(f"DEBUG: Found {len(existing_units_dict)} existing units in database")
        except Exception as fetch_error:
            print(f"DEBUG: Error fetching existing units: {fetch_error}")
            existing_units_dict = {}

        # Process units in smaller batches to avoid timeout
        # Smaller batch size and use bulk_create for better performance
        batch_size = 10  # Reduced from 50 to 10

        for batch_start in range(0, len(units_data), batch_size):
            batch_end = min(batch_start + batch_size, len(units_data))
            batch = units_data[batch_start:batch_end]

            print(f"DEBUG: Processing batch {batch_start//batch_size + 1} ({batch_start+1}-{batch_end} of {len(units_data)})")

            # Separate units to create vs update
            units_to_create = []
            units_to_update = []

            for unit_data in batch:
                try:
                    # Validate all string fields are properly encoded before saving
                    for field_name, field_value in unit_data.items():
                        if isinstance(field_value, str):
                            try:
                                # Ensure the string is valid UTF-8
                                field_value.encode('utf-8')
                            except UnicodeEncodeError:
                                # Clean the string by removing problematic characters
                                unit_data[field_name] = field_value.encode('utf-8', errors='ignore').decode('utf-8', errors='ignore')

                    # Check if unit already exists using pre-fetched dictionary
                    unit_number = unit_data['number']
                    existing_unit_id = existing_units_dict.get(unit_number)

                    if existing_unit_id:
                        units_to_update.append((existing_unit_id, unit_data))
                    else:
                        units_to_create.append(unit_data)

                except Exception as e:
                    error_msg = str(e)
                    unit_number = unit_data.get('number', 'unknown')
                    print(f"DEBUG: Error preparing unit {unit_number}: {error_msg}")
                    save_errors.append(f"Error preparing unit {unit_number}: {error_msg}")

            # Process updates without transaction to avoid commit timeout
            for existing_unit_id, unit_data in units_to_update:
                try:
                    existing_unit = Unit.objects.get(id=existing_unit_id)
                    for field, value in unit_data.items():
                        if field != 'building':  # Don't update building
                            setattr(existing_unit, field, value)
                    existing_unit.save()
                    created_units.append(existing_unit)
                    update_count += 1
                    print(f"DEBUG: Updated unit {unit_data['number']}")
                except Exception as e:
                    error_msg = str(e)
                    unit_number = unit_data.get('number', 'unknown')
                    print(f"DEBUG: Error updating unit {unit_number}: {error_msg}")
                    save_errors.append(f"Error updating unit {unit_number}: {error_msg}")

            # Process creates using bulk_create for maximum speed
            # Prepare all units first
            units_to_bulk_create = []
            for unit_data in units_to_create:
                try:
                    # Round decimal fields to proper precision to avoid validation errors
                    if 'area' in unit_data and unit_data['area'] is not None:
                        unit_data['area'] = round(float(unit_data['area']), 2)
                    if 'ideal_fraction' in unit_data and unit_data['ideal_fraction'] is not None:
                        unit_data['ideal_fraction'] = round(float(unit_data['ideal_fraction']), 6)

                    # Create unit object (not saved yet)
                    unit = Unit(**unit_data)
                    units_to_bulk_create.append(unit)
                except Exception as e:
                    error_msg = str(e)
                    unit_number = unit_data.get('number', 'unknown')
                    print(f"DEBUG: Error preparing unit {unit_number}: {error_msg}")
                    save_errors.append(f"Error preparing unit {unit_number}: {error_msg}")

            # Bulk create all units in one database operation
            if units_to_bulk_create:
                try:
                    print(f"DEBUG: Bulk creating {len(units_to_bulk_create)} units")
                    # Use bulk_create with batch_size to avoid memory issues
                    created_batch = Unit.objects.bulk_create(units_to_bulk_create, batch_size=10)
                    created_units.extend(created_batch)
                    create_count += len(created_batch)
                    print(f"DEBUG: Successfully bulk created {len(created_batch)} units")

                    # Update dictionary with new IDs
                    for unit in created_batch:
                        existing_units_dict[unit.number] = unit.id
                except Exception as e:
                    error_msg = str(e)
                    print(f"DEBUG: Error in bulk_create: {error_msg}")

                    # Fallback to individual saves if bulk_create fails
                    print(f"DEBUG: Falling back to individual saves")
                    for unit in units_to_bulk_create:
                        try:
                            unit.save()
                            created_units.append(unit)
                            create_count += 1
                            existing_units_dict[unit.number] = unit.id
                            print(f"DEBUG: Created unit {unit.number}")
                        except Exception as save_error:
                            error_msg = str(save_error)
                            print(f"DEBUG: Error creating unit {unit.number}: {error_msg}")
                            if "value too long for type character varying" in error_msg:
                                save_errors.append(f"Error saving unit {unit.number}: One or more fields exceed maximum length limits.")
                            elif "UnicodeDecodeError" in error_msg or "UnicodeEncodeError" in error_msg or "codec" in error_msg:
                                save_errors.append(f"Error saving unit {unit.number}: Invalid character encoding in data.")
                            elif "unique constraint" in error_msg.lower() or "duplicate key" in error_msg.lower():
                                save_errors.append(f"Error saving unit {unit.number}: Unit already exists.")
                            else:
                                save_errors.append(f"Error saving unit {unit.number}: {error_msg}")

        # Return response
        response_data = {
            'message': f'Successfully processed {len(created_units)} units',
            'summary': {
                'total_processed': len(created_units),
                'created': create_count,
                'updated': update_count
            }
        }

        if save_errors:
            response_data['warnings'] = save_errors

        if errors:
            response_data['validation_warnings'] = errors

        # Add unit data for frontend
        if created_units:
            try:
                serializer = UnitDetailSerializer(created_units, many=True)
                response_data['units'] = serializer.data
            except Exception as serializer_error:
                # Log serializer error but don't fail the entire request
                response_data['serializer_warning'] = f'Error serializing units: {str(serializer_error)}'
                # Return basic unit info as fallback
                response_data['units'] = []
                for unit in created_units:
                    try:
                        unit_data = {
                            'id': unit.id,
                            'number': unit.number,
                            'building_name': unit.building.building_name,
                            'building_id': unit.building.id,
                            'tower_id': unit.tower.id if unit.tower else None,
                            'tower_name': unit.tower.name if unit.tower else None,
                            'status': unit.status,
                            'area': float(unit.area),
                            'ideal_fraction': float(unit.ideal_fraction),
                            'floor': unit.floor,
                            'identification': unit.identification,
                            'owner': unit.owner,
                            'owner_phone': unit.owner_phone,
                            'parking_spaces': unit.parking_spaces,
                            'key_delivery': unit.key_delivery,
                            'deposit_location': unit.deposit_location
                        }
                        response_data['units'].append(unit_data)
                    except Exception as unit_error:
                        # Skip problematic units but log the issue
                        if 'unit_errors' not in response_data:
                            response_data['unit_errors'] = []
                        response_data['unit_errors'].append(f'Error serializing unit {unit.number}: {str(unit_error)}')

        status_code = status.HTTP_201_CREATED if created_units else status.HTTP_400_BAD_REQUEST
        return Response(response_data, status=status_code)

    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        print(f"CRITICAL ERROR in Excel import: {str(e)}")
        print(f"Error type: {type(e).__name__}")
        print(f"Full traceback: {error_traceback}")

        return Response({
            'error': f'Unexpected error processing Excel file: {str(e)}',
            'type': type(e).__name__,
            'traceback': error_traceback if request.user.is_superuser else None
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def debug_unit(request, id):
    """Debug endpoint to check unit existence and access rights."""
    try:
        # Check if unit exists
        unit = Unit.objects.get(id=id)

        # Get building info
        building = unit.building

        return Response({
            'unit_found': True,
            'unit_id': unit.id,
            'unit_number': unit.number,
            'building_id': building.id,
            'building_name': building.building_name,
            'building_created_by': str(building.created_by),
            'building_created_by_id': building.created_by.id,
            'request_user': str(request.user),
            'request_user_id': request.user.id,
            'has_access': building.created_by == request.user,
            'unit_details': {
                'floor': unit.floor,
                'area': float(unit.area),
                'status': unit.status,
                'owner': unit.owner,
                'tower': unit.tower.name if unit.tower else None
            }
        }, status=status.HTTP_200_OK)

    except Unit.DoesNotExist:
        return Response({
            'unit_found': False,
            'unit_id': id,
            'error': f'Unit with ID {id} does not exist',
            'request_user': str(request.user),
            'request_user_id': request.user.id
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({
            'error': f'Unexpected error: {str(e)}',
            'unit_id': id,
            'request_user': str(request.user)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def test_import_endpoint(request, id):
    """Simple test endpoint to verify POST requests work for this building."""
    try:
        print(f"TEST DEBUG: Building ID: {id}, User: {request.user}")
        print(f"TEST DEBUG: Request method: {request.method}")
        print(f"TEST DEBUG: Request FILES: {list(request.FILES.keys())}")
        print(f"TEST DEBUG: Request DATA: {list(request.data.keys())}")

        # Check building exists
        building = Building.objects.get(id=id, created_by=request.user)

        return Response({
            'success': True,
            'message': f'Test endpoint working for building {building.building_name}',
            'building_id': building.id,
            'user': str(request.user),
            'files_received': list(request.FILES.keys()),
            'data_received': list(request.data.keys())
        }, status=status.HTTP_200_OK)

    except Building.DoesNotExist:
        return Response({
            'error': 'Building not found or access denied',
            'building_id': id,
            'user': str(request.user)
        }, status=status.HTTP_404_NOT_FOUND)

    except Exception as e:
        import traceback
        return Response({
            'error': f'Test endpoint error: {str(e)}',
            'type': type(e).__name__,
            'traceback': traceback.format_exc()
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def simple_import_test(request, id):
    """Minimal import test to isolate the issue."""
    try:
        print(f"SIMPLE TEST: Starting import test for building {id}")

        # Just return success with basic info
        return Response({
            'success': True,
            'message': 'Simple import test successful',
            'building_id': id,
            'method': request.method,
            'user': str(request.user)
        }, status=status.HTTP_200_OK)

    except Exception as e:
        print(f"SIMPLE TEST ERROR: {str(e)}")
        return Response({
            'error': f'Simple test failed: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)